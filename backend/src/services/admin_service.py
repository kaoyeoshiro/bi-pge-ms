"""Serviço administrativo para gestão de roles e importação de dados."""

import hashlib
import io
import logging
import time
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy import delete, distinct, func, select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from src.config import settings
from src.domain.models import (
    HiddenProcuradorProducao,
    PecaElaborada,
    PecaFinalizada,
    Pendencia,
    ProcessoNovo,
    ProcuradorLotacao,
    TABLE_MODEL_MAP,
    UserRole,
)
from src.domain.schemas import HiddenProcuradorCreate, HiddenProcuradorResponse, HiddenProcuradorUpdate
from src.services.normalization import normalize_chefia_expr
from src.services.cache import clear_all_caches

logger = logging.getLogger(__name__)

# Tokens ativos em memória (sessão simples, sem JWT)
_active_tokens: set[str] = set()

# Mapeamento de colunas do Excel (cabeçalho) para colunas do banco
COLUMN_MAPS: dict[str, dict[str, str]] = {
    "processos_novos": {
        "Chefia": "chefia",
        "Data": "data",
        "Código do Processo": "codigo_processo",
        "Número do Processo": "numero_processo",
        "Número Formatado": "numero_formatado",
        "Procurador": "procurador",
    },
    "pecas_elaboradas": {
        "Chefia": "chefia",
        "Data": "data",
        "Usuário Criação": "usuario_criacao",
        "Categoria": "categoria",
        "Modelo": "modelo",
        "Número do Processo": "numero_processo",
        "Número Formatado": "numero_formatado",
        "Procurador": "procurador",
    },
    "pendencias": {
        "Chefia": "chefia",
        "Data": "data",
        "Número do Processo": "numero_processo",
        "Número Formatado": "numero_formatado",
        "Área": "area",
        "Procurador": "procurador",
        "Usuário Cumpridor Pendência": "usuario_cumpridor_pendencia",
        "Categoria": "categoria",
        "Categoria Pendência": "categoria_pendencia",
        "CDPENDENCIA": "cd_pendencia",
    },
    "pecas_finalizadas": {
        "Chefia": "chefia",
        "Data de Finalização": "data_finalizacao",
        "Usuário Finalização": "usuario_finalizacao",
        "Categoria": "categoria",
        "Modelo": "modelo",
        "Número do Processo": "numero_processo",
        "Número Formatado": "numero_formatado",
        "Procurador": "procurador",
    },
}

# Colunas de data por tabela
DATE_COLUMNS = {"data", "data_finalizacao"}

# Colunas numéricas
NUMERIC_COLUMNS = {"numero_processo", "cd_pendencia"}


class AdminAuthService:
    """Gerencia autenticação simples do painel admin."""

    @staticmethod
    def verify_password(password: str) -> bool:
        """Verifica se a senha fornecida é válida."""
        return password == settings.admin_password

    @staticmethod
    def generate_token() -> str:
        """Gera um token simples baseado na senha e timestamp."""
        raw = f"{settings.admin_password}:{time.time()}"
        token = hashlib.sha256(raw.encode()).hexdigest()
        _active_tokens.add(token)
        return token

    @staticmethod
    def verify_token(token: str) -> bool:
        """Verifica se o token está ativo."""
        return token in _active_tokens

    @staticmethod
    def revoke_token(token: str) -> None:
        """Revoga um token ativo."""
        _active_tokens.discard(token)


class UserRoleService:
    """Gerencia classificação de usuários (procurador/assessor)."""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def get_all_users(
        self,
        search: str | None = None,
        role: str | None = None,
    ) -> list[dict]:
        """Retorna todos os usuários com seus roles, em ordem alfabética."""
        stmt = select(UserRole).order_by(UserRole.name)
        if search:
            stmt = stmt.where(UserRole.name.ilike(f"%{search}%"))
        if role:
            stmt = stmt.where(UserRole.role == role)
        result = await self.session.execute(stmt)
        return [
            {
                "name": row.name,
                "role": row.role,
                "carga_reduzida": row.carga_reduzida or False,
            }
            for row in result.scalars().all()
        ]

    async def update_carga_reduzida(self, name: str, carga_reduzida: bool) -> dict:
        """Atualiza o flag de carga reduzida de um usuário."""
        stmt = (
            pg_insert(UserRole)
            .values(
                name=name,
                role="procurador",
                carga_reduzida=carga_reduzida,
                updated_at=func.now(),
            )
            .on_conflict_do_update(
                index_elements=[UserRole.name],
                set_={"carga_reduzida": carga_reduzida, "updated_at": func.now()},
            )
            .returning(UserRole.name, UserRole.carga_reduzida)
        )
        result = await self.session.execute(stmt)
        await self.session.commit()
        row = result.one()
        clear_all_caches()
        return {"name": row.name, "carga_reduzida": row.carga_reduzida}

    async def get_carga_reduzida_names(self) -> list[str]:
        """Retorna nomes de usuários com carga reduzida ativa."""
        stmt = (
            select(UserRole.name)
            .where(UserRole.carga_reduzida.is_(True))
            .order_by(UserRole.name)
        )
        result = await self.session.execute(stmt)
        return [row[0] for row in result.all()]

    async def update_role(self, name: str, role: str) -> dict:
        """Atualiza ou insere o role de um usuário."""
        stmt = (
            pg_insert(UserRole)
            .values(name=name, role=role, updated_at=func.now())
            .on_conflict_do_update(
                index_elements=[UserRole.name],
                set_={"role": role, "updated_at": func.now()},
            )
            .returning(UserRole.name, UserRole.role)
        )
        result = await self.session.execute(stmt)
        await self.session.commit()
        row = result.one()
        clear_all_caches()
        return {"name": row.name, "role": row.role}

    async def update_roles_bulk(self, users: list[dict]) -> int:
        """Atualiza roles de múltiplos usuários em lote."""
        count = 0
        for user in users:
            stmt = (
                pg_insert(UserRole)
                .values(name=user["name"], role=user["role"], updated_at=func.now())
                .on_conflict_do_update(
                    index_elements=[UserRole.name],
                    set_={"role": user["role"], "updated_at": func.now()},
                )
            )
            await self.session.execute(stmt)
            count += 1
        await self.session.commit()
        clear_all_caches()
        return count

    async def populate_initial_roles(self) -> dict[str, int]:
        """Popula a tabela user_roles com nomes extraídos dos dados existentes.

        Procuradores são identificados pela coluna 'procurador'.
        Assessores são os que aparecem em usuario_criacao/usuario_finalizacao
        mas NÃO na coluna procurador.
        """
        # Criar tabela se não existir
        await self.session.execute(text("""
            CREATE TABLE IF NOT EXISTS user_roles (
                id SERIAL PRIMARY KEY,
                name VARCHAR(300) UNIQUE NOT NULL,
                role VARCHAR(20) NOT NULL CHECK (role IN ('procurador', 'assessor')),
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """))

        # Coletar todos os procuradores distintos
        proc_names: set[str] = set()
        for model in [ProcessoNovo, PecaElaborada, PecaFinalizada, Pendencia]:
            stmt = (
                select(distinct(model.procurador))
                .where(model.procurador.isnot(None))
                .where(model.procurador != "")
            )
            result = await self.session.execute(stmt)
            proc_names.update(str(row[0]) for row in result.all())

        # Coletar todos os assessores (usuario_criacao e usuario_finalizacao)
        assessor_names: set[str] = set()
        for model, col_name in [
            (PecaElaborada, "usuario_criacao"),
            (PecaFinalizada, "usuario_finalizacao"),
            (Pendencia, "usuario_cumpridor_pendencia"),
        ]:
            col = getattr(model, col_name)
            stmt = (
                select(distinct(col))
                .where(col.isnot(None))
                .where(col != "")
            )
            result = await self.session.execute(stmt)
            assessor_names.update(str(row[0]) for row in result.all())

        # Assessores = quem aparece como usuário mas não como procurador
        only_assessores = assessor_names - proc_names

        # Limpar tabela e reinserir
        await self.session.execute(delete(UserRole))

        inserted_proc = 0
        inserted_assessor = 0

        for name in sorted(proc_names):
            await self.session.execute(
                pg_insert(UserRole)
                .values(name=name, role="procurador", carga_reduzida=False)
                .on_conflict_do_nothing()
            )
            inserted_proc += 1

        for name in sorted(only_assessores):
            await self.session.execute(
                pg_insert(UserRole)
                .values(name=name, role="assessor", carga_reduzida=False)
                .on_conflict_do_nothing()
            )
            inserted_assessor += 1

        await self.session.commit()
        clear_all_caches()

        logger.info(
            f"Roles populados: {inserted_proc} procuradores, "
            f"{inserted_assessor} assessores"
        )
        return {"procuradores": inserted_proc, "assessores": inserted_assessor}

    async def get_role_counts(self) -> dict[str, int]:
        """Retorna contagem de procuradores e assessores."""
        stmt = select(UserRole.role, func.count()).group_by(UserRole.role)
        result = await self.session.execute(stmt)
        counts = {"procurador": 0, "assessor": 0}
        for role, total in result.all():
            counts[role] = total
        return counts


class LotacaoService:
    """Gerencia lotação de procuradores em chefias."""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def get_all_lotacoes(
        self, search: str | None = None
    ) -> list[dict]:
        """Retorna lotações agrupadas por procurador."""
        stmt = select(ProcuradorLotacao).order_by(
            ProcuradorLotacao.procurador, ProcuradorLotacao.chefia
        )
        if search:
            stmt = stmt.where(ProcuradorLotacao.procurador.ilike(f"%{search}%"))
        result = await self.session.execute(stmt)
        rows = result.scalars().all()

        # Agrupar por procurador
        agrupado: dict[str, list[str]] = {}
        for row in rows:
            agrupado.setdefault(row.procurador, []).append(row.chefia)

        return [
            {"procurador": proc, "chefias": chefias}
            for proc, chefias in sorted(agrupado.items())
        ]

    async def set_lotacoes(self, procurador: str, chefias: list[str]) -> dict:
        """Define as chefias de um procurador (substitui todas as anteriores)."""
        # Remover lotações existentes
        await self.session.execute(
            delete(ProcuradorLotacao).where(
                ProcuradorLotacao.procurador == procurador
            )
        )

        # Inserir novas
        for chefia in chefias:
            await self.session.execute(
                pg_insert(ProcuradorLotacao)
                .values(procurador=procurador, chefia=chefia)
                .on_conflict_do_nothing()
            )

        await self.session.commit()
        return {"procurador": procurador, "chefias": chefias}

    async def get_chefias_disponiveis(self) -> list[str]:
        """Retorna chefias distintas normalizadas (PS unificada) das 4 tabelas."""
        all_chefias: set[str] = set()

        for model in [ProcessoNovo, PecaElaborada, PecaFinalizada, Pendencia]:
            normalized = normalize_chefia_expr(model.chefia)
            stmt = (
                select(distinct(normalized))
                .where(model.chefia.isnot(None))
                .where(model.chefia != "")
            )
            result = await self.session.execute(stmt)
            all_chefias.update(str(row[0]) for row in result.all())

        return sorted(all_chefias)


class ExcelImportService:
    """Gerencia importação de planilhas Excel para as tabelas do BI."""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def import_excel(
        self,
        file_content: bytes,
        tabela: str,
        modo: str,
    ) -> dict:
        """Importa dados de um arquivo Excel para a tabela especificada.

        Args:
            file_content: Conteúdo binário do arquivo Excel.
            tabela: Nome da tabela destino.
            modo: 'substituir' (TRUNCATE + INSERT) ou 'adicionar' (INSERT).

        Returns:
            Dicionário com contagem de linhas importadas e total.
        """
        if tabela not in COLUMN_MAPS:
            raise ValueError(f"Tabela inválida: {tabela}")

        column_map = COLUMN_MAPS[tabela]
        model = TABLE_MODEL_MAP[tabela]

        # Ler Excel com openpyxl
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Planilha não contém dados.")

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise ValueError("Planilha vazia ou sem dados além do cabeçalho.")

        # Mapear cabeçalho
        header = [str(cell).strip() if cell else "" for cell in rows[0]]
        col_indices: dict[str, int] = {}
        for excel_col, pg_col in column_map.items():
            try:
                col_indices[pg_col] = header.index(excel_col)
            except ValueError:
                raise ValueError(
                    f"Coluna '{excel_col}' não encontrada no Excel. "
                    f"Colunas disponíveis: {header}"
                )

        data_rows = rows[1:]

        # Se modo substituir, limpar tabela
        if modo == "substituir":
            await self.session.execute(
                text(f"TRUNCATE TABLE {tabela} RESTART IDENTITY CASCADE")
            )
            logger.info(f"Tabela {tabela} truncada para substituição.")

        # Inserir em chunks
        chunk_size = 5000
        total_imported = 0

        for i in range(0, len(data_rows), chunk_size):
            chunk = data_rows[i : i + chunk_size]
            records = []

            for row in chunk:
                record = {}
                for pg_col, idx in col_indices.items():
                    value = row[idx] if idx < len(row) else None
                    record[pg_col] = self._convert_value(pg_col, value)
                records.append(record)

            if records:
                await self.session.execute(
                    model.__table__.insert(),
                    records,
                )
                total_imported += len(records)

            logger.info(f"  {tabela}: {total_imported} linhas importadas...")

        await self.session.commit()

        # Recriar índices se modo substituir
        if modo == "substituir":
            await self._recreate_indexes(tabela)

        # Limpar caches
        clear_all_caches()

        # Contar total na tabela
        stmt = select(func.count()).select_from(model)
        result = await self.session.execute(stmt)
        total_tabela = result.scalar() or 0

        wb.close()

        logger.info(
            f"Importação concluída: {total_imported} linhas para {tabela}. "
            f"Total na tabela: {total_tabela}"
        )
        return {
            "linhas_importadas": total_imported,
            "linhas_total_tabela": total_tabela,
        }

    @staticmethod
    def _convert_value(column: str, value):
        """Converte valor da célula para o tipo correto da coluna."""
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return None

        if column in DATE_COLUMNS:
            if isinstance(value, datetime):
                return value
            try:
                return datetime.fromisoformat(str(value))
            except (ValueError, TypeError):
                return None

        if column in NUMERIC_COLUMNS:
            try:
                return int(float(str(value)))
            except (ValueError, TypeError):
                return None

        return str(value).strip() if value else None

    async def _recreate_indexes(self, tabela: str) -> None:
        """Recria índices para a tabela após TRUNCATE."""
        index_map = {
            "processos_novos": [
                "CREATE INDEX IF NOT EXISTS idx_processos_novos_numero ON processos_novos(numero_processo)",
                "CREATE INDEX IF NOT EXISTS idx_processos_novos_chefia ON processos_novos(chefia)",
                "CREATE INDEX IF NOT EXISTS idx_processos_novos_procurador ON processos_novos(procurador)",
                "CREATE INDEX IF NOT EXISTS idx_processos_novos_data ON processos_novos(data)",
            ],
            "pecas_elaboradas": [
                "CREATE INDEX IF NOT EXISTS idx_pecas_elaboradas_numero ON pecas_elaboradas(numero_processo)",
                "CREATE INDEX IF NOT EXISTS idx_pecas_elaboradas_chefia ON pecas_elaboradas(chefia)",
                "CREATE INDEX IF NOT EXISTS idx_pecas_elaboradas_procurador ON pecas_elaboradas(procurador)",
                "CREATE INDEX IF NOT EXISTS idx_pecas_elaboradas_data ON pecas_elaboradas(data)",
                "CREATE INDEX IF NOT EXISTS idx_pecas_elaboradas_categoria ON pecas_elaboradas(categoria)",
            ],
            "pendencias": [
                "CREATE INDEX IF NOT EXISTS idx_pendencias_numero ON pendencias(numero_processo)",
                "CREATE INDEX IF NOT EXISTS idx_pendencias_chefia ON pendencias(chefia)",
                "CREATE INDEX IF NOT EXISTS idx_pendencias_procurador ON pendencias(procurador)",
                "CREATE INDEX IF NOT EXISTS idx_pendencias_data ON pendencias(data)",
                "CREATE INDEX IF NOT EXISTS idx_pendencias_area ON pendencias(area)",
                "CREATE INDEX IF NOT EXISTS idx_pendencias_categoria ON pendencias(categoria)",
            ],
            "pecas_finalizadas": [
                "CREATE INDEX IF NOT EXISTS idx_pecas_finalizadas_numero ON pecas_finalizadas(numero_processo)",
                "CREATE INDEX IF NOT EXISTS idx_pecas_finalizadas_chefia ON pecas_finalizadas(chefia)",
                "CREATE INDEX IF NOT EXISTS idx_pecas_finalizadas_procurador ON pecas_finalizadas(procurador)",
                "CREATE INDEX IF NOT EXISTS idx_pecas_finalizadas_data ON pecas_finalizadas(data_finalizacao)",
                "CREATE INDEX IF NOT EXISTS idx_pecas_finalizadas_categoria ON pecas_finalizadas(categoria)",
            ],
        }
        for idx_sql in index_map.get(tabela, []):
            await self.session.execute(text(idx_sql))
        await self.session.commit()


class TableStatsService:
    """Fornece estatísticas das tabelas do BI."""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def get_stats(self) -> list[dict]:
        """Retorna contagem de linhas de cada tabela."""
        stats = []
        for table_name, model in TABLE_MODEL_MAP.items():
            stmt = select(func.count()).select_from(model)
            result = await self.session.execute(stmt)
            count = result.scalar() or 0
            stats.append({"tabela": table_name, "total": count})
        return stats


class HiddenProducaoService:
    """Gerencia regras de ocultação temporária de produção de procuradores."""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def list_rules(
        self, only_active: bool = True
    ) -> list[HiddenProcuradorResponse]:
        """Lista regras de ocultação, opcionalmente filtrando apenas ativas."""
        stmt = select(HiddenProcuradorProducao).order_by(
            HiddenProcuradorProducao.created_at.desc()
        )
        if only_active:
            stmt = stmt.where(HiddenProcuradorProducao.is_active.is_(True))
        result = await self.session.execute(stmt)
        return [
            HiddenProcuradorResponse(
                id=row.id,
                procurador_name=row.procurador_name,
                chefia=row.chefia,
                start_date=row.start_date,
                end_date=row.end_date,
                is_active=row.is_active,
                reason=row.reason,
                created_by=row.created_by,
                created_at=row.created_at,
                updated_at=row.updated_at,
            )
            for row in result.scalars().all()
        ]

    async def create_rule(
        self, data: HiddenProcuradorCreate
    ) -> HiddenProcuradorResponse:
        """Cria nova regra de ocultação.

        Valida duplicatas (mesmo procurador + chefia com período sobreposto).
        """
        # Verificar sobreposição de período
        overlap_stmt = (
            select(HiddenProcuradorProducao)
            .where(
                HiddenProcuradorProducao.procurador_name == data.procurador_name,
                HiddenProcuradorProducao.is_active.is_(True),
                HiddenProcuradorProducao.start_date <= data.end_date,
                HiddenProcuradorProducao.end_date >= data.start_date,
            )
        )
        if data.chefia is None:
            overlap_stmt = overlap_stmt.where(
                HiddenProcuradorProducao.chefia.is_(None)
            )
        else:
            overlap_stmt = overlap_stmt.where(
                HiddenProcuradorProducao.chefia == data.chefia
            )
        result = await self.session.execute(overlap_stmt)
        if result.scalars().first():
            raise ValueError(
                "Já existe regra ativa com período sobreposto para este procurador/chefia."
            )

        rule = HiddenProcuradorProducao(
            procurador_name=data.procurador_name,
            chefia=data.chefia,
            start_date=data.start_date,
            end_date=data.end_date,
            reason=data.reason,
        )
        self.session.add(rule)
        await self.session.commit()
        await self.session.refresh(rule)
        clear_all_caches()

        return HiddenProcuradorResponse(
            id=rule.id,
            procurador_name=rule.procurador_name,
            chefia=rule.chefia,
            start_date=rule.start_date,
            end_date=rule.end_date,
            is_active=rule.is_active,
            reason=rule.reason,
            created_by=rule.created_by,
            created_at=rule.created_at,
            updated_at=rule.updated_at,
        )

    async def update_rule(
        self, rule_id: int, data: HiddenProcuradorUpdate
    ) -> HiddenProcuradorResponse:
        """Atualiza uma regra de ocultação existente."""
        stmt = select(HiddenProcuradorProducao).where(
            HiddenProcuradorProducao.id == rule_id
        )
        result = await self.session.execute(stmt)
        rule = result.scalars().first()
        if not rule:
            raise ValueError(f"Regra #{rule_id} não encontrada.")

        if data.start_date is not None:
            rule.start_date = data.start_date
        if data.end_date is not None:
            rule.end_date = data.end_date
        if data.is_active is not None:
            rule.is_active = data.is_active
        if data.reason is not None:
            rule.reason = data.reason

        # Validar datas após update
        if rule.start_date > rule.end_date:
            raise ValueError("start_date deve ser anterior ou igual a end_date.")

        await self.session.commit()
        await self.session.refresh(rule)
        clear_all_caches()

        return HiddenProcuradorResponse(
            id=rule.id,
            procurador_name=rule.procurador_name,
            chefia=rule.chefia,
            start_date=rule.start_date,
            end_date=rule.end_date,
            is_active=rule.is_active,
            reason=rule.reason,
            created_by=rule.created_by,
            created_at=rule.created_at,
            updated_at=rule.updated_at,
        )

    async def delete_rule(self, rule_id: int) -> None:
        """Remove uma regra de ocultação permanentemente."""
        stmt = select(HiddenProcuradorProducao).where(
            HiddenProcuradorProducao.id == rule_id
        )
        result = await self.session.execute(stmt)
        rule = result.scalars().first()
        if not rule:
            raise ValueError(f"Regra #{rule_id} não encontrada.")

        await self.session.delete(rule)
        await self.session.commit()
        clear_all_caches()
