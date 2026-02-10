"""Serviço para geração de relatório Excel de assuntos por chefia."""

import io
import logging
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.domain.filters import GlobalFilters
from src.domain.models import (
    Assunto,
    ProcessoAssunto,
    ProcessoNovo,
    TABLE_MODEL_MAP,
)
from src.repositories.base_repository import BaseRepository

logger = logging.getLogger(__name__)

# Estilos do Excel
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(bold=True, size=13, color="1F4E79")
PERCENT_FORMAT = "0.0%"
NUMBER_FORMAT = "#,##0"


@dataclass
class AssuntoReportRow:
    """Linha do relatório com dados de um assunto na hierarquia."""

    codigo: int
    nome: str
    nivel: int
    path_names: list[str] = field(default_factory=list)
    path_codigos: list[int] = field(default_factory=list)
    total: int = 0


class AssuntoReportService:
    """Gera relatório Excel completo de assuntos hierárquicos de uma chefia."""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def generate_excel(self, chefia: str, filters: GlobalFilters) -> bytes:
        """Gera relatório Excel completo com todos os assuntos da chefia.

        Retorna bytes do arquivo .xlsx pronto para download.
        """
        rows = await self._collect_data(chefia, filters)
        return self._build_workbook(rows, chefia, filters)

    async def _collect_data(
        self, chefia: str, filters: GlobalFilters
    ) -> list[AssuntoReportRow]:
        """Coleta contagens de assuntos e agrega bottom-up na hierarquia.

        Fase 1: conta processos distintos por assunto folha (código).
        Fase 2: carrega árvore de assuntos e acumula totais subindo.
        """
        # Fase 1 — Contagem por assunto (código) em processos_novos
        model = ProcessoNovo
        repo = BaseRepository(self.session, model)

        # Monta filtro com chefia preenchida
        f = GlobalFilters(
            anos=filters.anos,
            mes=filters.mes,
            data_inicio=filters.data_inicio,
            data_fim=filters.data_fim,
            chefia=[chefia],
        )

        stmt = (
            select(
                ProcessoAssunto.codigo_assunto.label("codigo"),
                func.count(func.distinct(model.id)).label("total"),
            )
            .select_from(model)
            .join(
                ProcessoAssunto,
                ProcessoAssunto.numero_processo == model.numero_processo,
            )
            .where(ProcessoAssunto.assunto_principal.is_(True))
            .group_by(ProcessoAssunto.codigo_assunto)
        )
        stmt = repo._apply_global_filters(stmt, f)

        result = await self.session.execute(stmt)
        leaf_counts: dict[int, int] = {
            row.codigo: row.total for row in result.all()
        }

        if not leaf_counts:
            return []

        # Fase 2 — Carregar todos os assuntos e agregar bottom-up
        all_assuntos_result = await self.session.execute(
            select(Assunto.codigo, Assunto.codigo_pai, Assunto.nome, Assunto.nivel)
        )
        all_assuntos = all_assuntos_result.all()

        # Mapas auxiliares
        info: dict[int, tuple[str, int, int | None]] = {}  # codigo → (nome, nivel, pai)
        children_map: dict[int | None, list[int]] = {}  # pai → [filhos]

        for row in all_assuntos:
            info[row.codigo] = (row.nome or f"Código {row.codigo}", row.nivel or 0, row.codigo_pai)
            children_map.setdefault(row.codigo_pai, []).append(row.codigo)

        # Acumulação recursiva bottom-up (com cache)
        acumulados: dict[int, int] = {}

        def acumular(codigo: int) -> int:
            if codigo in acumulados:
                return acumulados[codigo]
            total = leaf_counts.get(codigo, 0)
            for filho in children_map.get(codigo, []):
                total += acumular(filho)
            acumulados[codigo] = total
            return total

        # Acumula a partir das raízes
        for cod in children_map.get(None, []):
            acumular(cod)
        # Garante que nós isolados (folhas sem pai conhecido) também sejam acumulados
        for cod in leaf_counts:
            if cod not in acumulados:
                acumular(cod)

        # Montar path (subindo até raiz) e criar rows
        def build_path(codigo: int) -> tuple[list[str], list[int]]:
            names: list[str] = []
            codigos: list[int] = []
            current = codigo
            while current is not None and current in info:
                nome, _, pai = info[current]
                names.append(nome)
                codigos.append(current)
                current = pai
            names.reverse()
            codigos.reverse()
            return names, codigos

        rows: list[AssuntoReportRow] = []
        for codigo, total in acumulados.items():
            if total <= 0:
                continue
            if codigo not in info:
                continue
            nome, nivel, _ = info[codigo]
            path_names, path_codigos = build_path(codigo)
            rows.append(AssuntoReportRow(
                codigo=codigo,
                nome=nome,
                nivel=nivel,
                path_names=path_names,
                path_codigos=path_codigos,
                total=total,
            ))

        rows.sort(key=lambda r: (r.nivel, -r.total))
        return rows

    def _build_workbook(
        self,
        rows: list[AssuntoReportRow],
        chefia: str,
        filters: GlobalFilters,
    ) -> bytes:
        """Monta workbook Excel com uma aba por nível hierárquico."""
        wb = Workbook()
        # Remove aba padrão
        wb.remove(wb.active)

        anos_str = ", ".join(str(a) for a in filters.anos) if filters.anos else "Todos"
        titulo_base = f"Relatório de Assuntos — {chefia} — Anos: {anos_str}"

        if not rows:
            ws = wb.create_sheet("Sem Dados")
            ws.append([titulo_base])
            ws["A1"].font = TITLE_FONT
            ws.append([])
            ws.append(["Nenhum assunto encontrado para os filtros selecionados."])
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        # Agrupar por nível
        by_level: dict[int, list[AssuntoReportRow]] = {}
        for row in rows:
            by_level.setdefault(row.nivel, []).append(row)

        # Mapa de acumulados para % do Pai
        acumulados_map: dict[int, int] = {r.codigo: r.total for r in rows}

        # Total geral (soma dos nível 1 = raízes)
        min_level = min(by_level.keys())
        total_geral = sum(r.total for r in by_level[min_level])

        for nivel in sorted(by_level.keys()):
            level_rows = sorted(by_level[nivel], key=lambda r: -r.total)

            if nivel == min_level:
                sheet_name = "Visão Geral"
            else:
                sheet_name = f"Nível {nivel}"

            ws = wb.create_sheet(sheet_name)

            # Título
            ws.append([titulo_base])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            ws["A1"].font = TITLE_FONT
            ws.append([])  # Linha em branco

            # Cabeçalhos
            headers: list[str] = []
            # Colunas de contexto: Nível 1 até Nível (N-1)
            context_levels = nivel - min_level
            for i in range(context_levels):
                headers.append(f"Nível {min_level + i}")
            headers.extend(["Assunto", "Processos"])
            if nivel > min_level:
                headers.append("% do Pai")
            headers.append("% do Total")

            header_row_idx = 3
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Dados
            for row_idx, row in enumerate(level_rows, header_row_idx + 1):
                col = 1

                # Colunas de contexto (ancestrais)
                for i in range(context_levels):
                    ancestor_idx = i  # índice no path
                    if ancestor_idx < len(row.path_names) - 1:
                        ws.cell(row=row_idx, column=col, value=row.path_names[ancestor_idx])
                    col += 1

                # Assunto (nome do nó atual)
                ws.cell(row=row_idx, column=col, value=row.nome)
                col += 1

                # Processos
                cell_proc = ws.cell(row=row_idx, column=col, value=row.total)
                cell_proc.number_format = NUMBER_FORMAT
                col += 1

                # % do Pai (apenas para níveis > mínimo)
                if nivel > min_level:
                    pai_codigo = row.path_codigos[-2] if len(row.path_codigos) >= 2 else None
                    pai_total = acumulados_map.get(pai_codigo, 0) if pai_codigo else 0
                    pct_pai = row.total / pai_total if pai_total > 0 else 0
                    cell_pct = ws.cell(row=row_idx, column=col, value=pct_pai)
                    cell_pct.number_format = PERCENT_FORMAT
                    col += 1

                # % do Total
                pct_total = row.total / total_geral if total_geral > 0 else 0
                cell_pct_total = ws.cell(row=row_idx, column=col, value=pct_total)
                cell_pct_total.number_format = PERCENT_FORMAT

            # Congelar cabeçalho
            ws.freeze_panes = f"A{header_row_idx + 1}"

            # Auto-width
            for col_idx in range(1, len(headers) + 1):
                max_len = len(str(headers[col_idx - 1]))
                for row_idx in range(header_row_idx + 1, header_row_idx + 1 + len(level_rows)):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val is not None:
                        cell_len = len(str(cell_val))
                        if cell_len > max_len:
                            max_len = cell_len
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
